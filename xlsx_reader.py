import openpyxl
from unidecode import unidecode
import csv
import logging
import pytomlpp
from dataclasses import dataclass,field
from datetime import date,timedelta
import mcc_reader
import xlsx2odf
import edt_matieres




_logger = logging.getLogger(__name__)


@dataclass
class Cours:

    cellule : str
    codeMatiere : str = ""  #code apogée
    nature : str = ""
    groupe : str = ""
    semaine : str = ""
    jour : date = None
    heure : str = ""
    duree : int = 0
    prof : str = ""
    note : str = ""   #ce qu'il y a entre crochet
    verificationGroupe : int = 0   #groupe écrit dans l'edt (ex: Gr1)
    numeroEdt : int = 0  #numéro écrit dans l'edt
    numero : int = 0  

    def csv(self):
        #Renvoie une liste avec les informations pour remplir le fichier des cours en csv
        return [self.codeMatiere, self.nature, self.groupe, self.jour.strftime("%d/%m/20%y"), self.semaine, self.heure, self.duree, "", self.prof, "", self.note]
    
    def __lt__(self, other):
        if self.jour == other.jour:
            return int(self.heure.split(":")[0]) < int(other.heure.split(":")[0])
        else:    
            return self.jour < other.jour


@dataclass
class Matiere:  

    code : str = ""
    nom : str = ""
    acronymes : set[str] = field(default_factory = lambda: set())
    numeroModule : str = ""
    semestre : int = 0   #le semetre est à 0 si la matière n'existe pas dans l'odf
    TD : set = field(default_factory = lambda: set())
    coursParTD : dict = field(default_factory = lambda: {}) # {nature : {TD : [] for TD in matiere.TD} for nature in naturesPossibles}
    
    def getMatiereAcronyme(self,acronyme):
        return acronyme in self.acronymes


class MatiereError(Exception):
    pass

class NatureError(Exception):
    pass

class MaquetteError(Exception):
    pass



def lecteurMaquette(nomFichierMaquette,naturesPossibles,listeTD,matieres):
    """
    Lit l'onglet maquette du fichier nomFichierMaquette et renvoie :
        - un dictionnaire des matières avec en clés les codes apogée et en valeurs les matières.
        - la liste des dépendances écrites dans la maquette
    """

    workbookMaquette = openpyxl.load_workbook(nomFichierMaquette, read_only = True) 
    pageMaquette = workbookMaquette["maquette"]

    colonneCode = None
    colonneMatiere = None
    colonneAcronyme = None
    colonneDependance = None

    for colonne,titre in enumerate(pageMaquette[1]):
        if titre.value != None:
            if unidecode(titre.value).lower() == "code apogee":
                colonneCode = colonne
            if unidecode(titre.value).lower() == "matiere":
                colonneMatiere = colonne
            if unidecode(titre.value).lower() == "acronyme":
                colonneAcronyme = colonne
            if unidecode(titre.value).lower() == "dependances":
                colonneDependance = colonne

    if None in (colonneCode,colonneMatiere,colonneAcronyme):
        raise MaquetteError

    dependances = []

    codeMatiere = None
    for ligne in pageMaquette.iter_rows(min_row = 2):

        if ligne != ():
            if ligne[colonneMatiere].value != None and ligne[colonneAcronyme].value != None:
                if ligne[colonneCode].value not in (None,"="):   
                    codeMatiere = ligne[colonneCode].value
                #sinon on garde le même code que la ligne du dessus

                if not codeMatiere in matieres.keys():  #code apogée qui n'existe pas dans l'odf
                    nomMatiere = unidecode(ligne[colonneMatiere].value.lower())
                    _logger.error("Le code apogée %s de %s n'existe pas dans l'offre de formation", codeMatiere, nomMatiere)
                    mat = Matiere(code = codeMatiere, nom = nomMatiere, TD = set(listeTD))
                    mat.coursParTD = {nature : {TD : [] for TD in mat.TD} for nature in naturesPossibles}
                    matieres[mat.code] = mat

                matiere = matieres[codeMatiere]
                if ligne[colonneMatiere].value != "=":
                    matiere.acronymes.add(unidecode(ligne[colonneMatiere].value.lower()))

                for acronyme in unidecode(ligne[colonneAcronyme].value.lower()).split(","):
                    if "@s" in acronyme and not matiere.semestre in (int(acronyme.split("@s")[colonneMatiere].strip()), 0):
                        _logger.warning("Maquette : %s : %s   ==>   Erreur de semestre dans cet acronyme", matiere.nom, acronyme)
                    acronyme = acronyme.split("@")[colonneCode].strip()
                    matiere.acronymes.add(acronyme)

                if colonneDependance != None:
                    if ligne[colonneDependance].value != None:
                        for texte in ligne[colonneDependance].value.split(","):
                            dependances.append((codeMatiere,texte.strip()))


    #vérification que plusieurs matières du même semestre n'est pas le même acronyme
    verifAcronymes = {1 : set(), 2 : set()}
    for matiere in matieres.values():
        copieAcronymes = set(matiere.acronymes)
        for acronyme in copieAcronymes:
            if matiere.semestre == 0: #semestre 0 pour les matières qui ne sont pas dans l'offre de formation 
                if acronyme in verifAcronymes[1] or acronyme in verifAcronymes[2]:
                    matiere.acronymes.discard(acronyme)
                    _logger.error("Maquette : %s : %s   ==>   Cet acronyme existe déja pour une matière du même semestre", matiere.nom, acronyme)
                else:
                    verifAcronymes[1] |= {acronyme}
                    verifAcronymes[2] |= {acronyme}
            else:
                if acronyme in verifAcronymes[matiere.semestre]:
                    matiere.acronymes.discard(acronyme)
                    _logger.error("Maquette : %s : %s   ==>   Cet acronyme existe déja pour une matière du même semestre", matiere.nom, acronyme)
                else:
                    verifAcronymes[matiere.semestre] |= {acronyme}


    workbookMaquette.close()
    return matieres, dependances


def lecteurAlias(nomFichierAlias):
    """
    Renvoie un dictionnaire avec en clés les alias et en valeurs le nom complet des profs
    """
    aliasProfs = {}
    with open(nomFichierAlias, encoding="utf8") as csvAlias:
        reader = csv.DictReader((row for row in csvAlias if len(row) > 2 and row[0] != '#'))
                            # XXX 2 for a single comma and newline.
        for row in reader:
            alias = unidecode(row["alias"].strip().lower())
            nom = unidecode(row["value"].strip().lower())
            aliasProfs[alias] = nom
    return aliasProfs


def generateurColonne(dispositionJour, listeTD):
    """
    Crée les dictionnaire colonneJour et colonneTD qui associe chaque colonne de l'edt à un jour et à un groupe de TD
    Ils peuvent varier en fonction des semaines (ex: le vendredi peut être sur 1 colonne ou sur 7 pour les 1SN)
    """
    colonneJour = {i : [] for i in range(5)}
    i = -1
    for colonne,jour in enumerate(dispositionJour + [None] * 6):
        if jour in ("LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI"):
            i += 1 
        colonneJour[i].append(colonne + 2)

    colonneTD = {}
    for colonnes in colonneJour.values():    
        if len(colonnes) == len(listeTD):
            for indice,colonne in enumerate(colonnes):
                TD = listeTD[indice]
                colonneTD[TD] = colonneTD.get(TD, []) + [colonne]

    return colonneTD, colonneJour



def analyserElement(element, cours, ignores, naturesPossibles, matieres, aliasProfs, semestre, semaine):
    """
    Vérifie si la chaîne de caractère element correspond à quelque chose de connu (matière, nature, prof, groupe, numéro)
    Si oui, on la rajoute dans le cours à la bonne variable
    Sinon, on la renvoie pour qu'elle aille dans les mots incompris
    """

    #l'élément est inutile
    if element in ignores:
        pass
    
    #l'élément est le numéro du cours
    elif element.isdigit():
        if cours.numeroEdt == 0:
            cours.numeroEdt = int(element)

    #l'élément est un groupe de TD
    elif element.startswith("gr") and len(element) == 3 and element[2].isdigit():
        cours.verificationGroupe = int(element[2])

    #l'élément est un enseignant
    elif element in aliasProfs:
        if cours.prof != "":
            cours.prof += " + "
        cours.prof += aliasProfs[element]

    else:
        #l'élément est une matiere
        if cours.codeMatiere == "":
            for matiere in matieres.values():
                if matiere.getMatiereAcronyme(element):
                    cours.codeMatiere = matiere.code
                    if matiere.semestre in (0,semestre):
                        return None
                            
            if cours.codeMatiere != "":
                _logger.warning("Semaine : %s, cellule : %s     ==>    Attention le cours de %s n'est pas dans le bon semestre", semaine, cours.cellule, matieres[cours.codeMatiere].nom)
                return None


        elementSansChiffre = "".join(lettre for lettre in element if not lettre.isdigit())

        #l'élément est une nature
        if len([nature for nature,acronymes in naturesPossibles.items() if elementSansChiffre in acronymes]) > 0:
            if cours.nature == "":
                cours.nature = [nature for nature,acronymes in naturesPossibles.items() if elementSansChiffre in acronymes][0]
            else:
                raise NatureError
            
            #il y avait un chiffre avec la nature qui correspond au numéro du cours
            if elementSansChiffre != element:
                cours.numeroEdt = int(element.replace(elementSansChiffre,""))


        #l'élément est inconnu
        else:
            return element


def mini(ensemble):
    """
    Renvoie le mot le plus cours d'un ensemble
    """
    mini = ""
    for mot in ensemble:
        if len(mot) < len(mini) or mini == "":
            mini = mot
    return mini


def dateSemaine(annee, semaine, jour):
    """
    Renvoie une date en fonction de l'année, de la semaine et du jour
    """
    ref = date(annee, 1, 4) # Le 4 janvier est toujours en semaine 1
    return ref + timedelta(weeks = semaine - 1, days = jour - ref.weekday())


def decodeurDependance(cours, codeMatiereOrigine, matieres, naturesPossibles):
    """
    Récupère la matière, la nature et le numéro d'un cours à partir du texte de la dépendance
    """

    if ":" in cours:
        acronyme = unidecode(cours.split(":")[0].strip().lower())
        if "@s" in acronyme:
            semestre = int(acronyme.split("@s")[1])
            acronyme = acronyme.split("@s")[0]
        else:
            semestre = 0

        codeMatiere = None
        for matiere in matieres.values():
            if codeMatiere == None and matiere.getMatiereAcronyme(acronyme):
                if semestre in (0, matiere.semestre):
                    codeMatiere = matiere.code

        if codeMatiere == None:
            raise MatiereError
        
        nature = cours.split(":")[1]

    else:
        codeMatiere = codeMatiereOrigine
        nature = cours

    nature = unidecode(nature.strip().lower())
    natureSansNumero = "".join(lettre for lettre in nature if not lettre.isdigit()).strip()
    numero = nature.replace(natureSansNumero,"").strip()
    if numero == "":
        numero = 1
    else:
        numero = int(numero)
    nature = [nature for nature,acronymes in naturesPossibles.items() if natureSansNumero in acronymes][0]

    return codeMatiere, nature, numero


def lecteurDependance(codeMatiereOrigine,texte,matieres,naturesPossibles):
    """
    Formate les dépendances pour qu'elles soient compréhensible par la fonction comparaison
    """
    
    if " <" in texte or "< " in texte:
        _logger.error("Maquette : %s : %s   ==>   Erreur de format de dépendance", matieres[codeMatiereOrigine].nom, texte)
        return None

    dependance = texte.replace(" ","<")

    try:
        if "<<" in dependance:
            cours = unidecode(dependance.split("<<")[0].strip().lower())
            codeMatiere, nature, numero = decodeurDependance(cours, codeMatiereOrigine, matieres, naturesPossibles)
            return [[codeMatiere,nature,numero,"avant"]]

        elif ">>" in dependance:
            cours = unidecode(dependance.split(">>")[0].strip().lower())
            codeMatiere, nature, numero = decodeurDependance(cours, codeMatiereOrigine, matieres, naturesPossibles)
            return [[codeMatiere,nature,numero,"apres"]]

        elif "<" in dependance:
            dependanceFormatee = []
            for cours in dependance.split("<"):
                if "(" in cours: 
                    cours = cours.removeprefix("(").removesuffix(")")
                    listeCours = []
                    for c in cours.split("|"):
                        codeMatiere, nature, numero = decodeurDependance(c, codeMatiereOrigine, matieres, naturesPossibles)
                        listeCours.append([codeMatiere,nature,numero])
                    dependanceFormatee.append(listeCours)
                else:
                    codeMatiere, nature, numero = decodeurDependance(cours, codeMatiereOrigine, matieres, naturesPossibles)
                    dependanceFormatee.append([codeMatiere,nature,numero])
            return dependanceFormatee
        
        else:
            _logger.error("Maquette : %s : %s   ==>   Erreur de format de dépendance", matieres[codeMatiereOrigine].nom, texte)
            return None

    except MatiereError:
        _logger.error("Maquette : %s : %s   ==>   Erreur d'acronyme : %s", matieres[codeMatiereOrigine].nom, texte, cours.split(":")[0])
        return None

    except IndexError:
        _logger.error("Maquette : %s : %s   ==>   Erreur de nature : %s", matieres[codeMatiereOrigine].nom, texte, cours)
        return None



def ecritureDependance(nomFichier,matieres):
    """
    Ecrit dans un fichier l'ensemble des dépendances de tous les cours de l'edt 
    """
    with open(nomFichier, "w", encoding="utf8") as fichierDependances:
        for matiere in matieres.values():
            listeCours = {}
            nombreSeance = 0

            for TD in matiere.TD:
                listeCours[TD] = []
                for cours in matiere.coursParTD.values():
                    listeCours[TD] += cours[TD]
                if len(listeCours[TD]) != 0:
                    listeCours[TD].sort()
                    nombreSeance = max(nombreSeance,len(listeCours[TD]))
                    
            texteDependances = [{} for _ in range(nombreSeance)]
            for TD in matiere.TD:
                for numero,cours in enumerate(listeCours[TD]):
                    texte = f"{cours.nature}{cours.numero}"
                    texteDependances[numero][texte] = texteDependances[numero].get(texte,[]) + [TD]

            fichierDependances.write(f"{matiere.code} {matiere.numeroModule} {matiere.nom} :")
            buffer = {}
            for seances in texteDependances:

                if len(seances) == 1 and len(buffer) == 0:
                    fichierDependances.write(" " + next(iter(seances.keys())))

                else:
                    for seance in seances:
                        buffer[seance] = buffer.get(seance,[]) + seances[seance]
                        buffer[seance].sort()

                    if len(buffer) > 1:
                        if all([len(TDs) == len(matiere.TD) for TDs in buffer.values()]):
                            texte = " ("
                            for seance in buffer:
                                texte += seance + "|"
                            texte = texte.removesuffix("|") + ")"
                            fichierDependances.write(texte)
                            buffer = {}
            
            if buffer != {}:
                fichierDependances.write(" + restant : " + str(buffer))
            fichierDependances.write("\n")


def getCours(seances,TD,matieres,dependanceMaquette):
    """
    A partir de seances (liste de séances représentées par : codeMatière, nature, numéro) et d'un groupe de TD,
    renvoie une liste des cours correspondant pour qu'ils puissent après être comparés
    """
    cours = []
    for indice,seance in enumerate(seances):
        if isinstance(seance[0],list):
            cours.append(getCours(seance,TD,matieres))
        else:
            codeMatiere = seance[0]
            nature = seance[1]
            numero = seance[2]
            if len(matieres[codeMatiere].coursParTD[nature][TD]) >= numero:
                cours.append(matieres[codeMatiere].coursParTD[nature][TD][numero - 1])
            else:
                _logger.error("Maquette : %s : %s   ==>   Erreur de numéro de séances : %s", matieres[dependanceMaquette[0]].nom, dependanceMaquette[1], dependanceMaquette[1].replace(" ","<").split("<")[indice])
                raise IndexError
    return cours


def avant(cours1,cours2):
    """
    Permet de comparer deux cours, un cours et une liste de cours ou deux listes de cours pour savoir si cours1 est entièrement avant cours2
    """
    if isinstance(cours1,Cours):
        if isinstance(cours2,Cours):
            return cours1 < cours2
        
        else: #cours2 est une liste de cours
            return all(cours1 < c2 for c2 in cours2)
        
    else: #cours1 est une liste de cours
        if isinstance(cours2,Cours):
            return all(c1 < cours2 for c1 in cours1)
        
        else: #cours2 est une liste de cours
            return all(c1 < c2 for c1 in cours1 for c2 in cours2)
                


def comparaison(seances,matieres,listeTD,dependanceMaquette,condition = None):
    """
    Permet de vérifier une dépendance entre les cours présent dans séances pour chaque groupe de TD concerné
    """
    TDEnCommun = set(listeTD)
    for seance in seances:
        if isinstance(seance[0],list):
            for s in seance:
                TDEnCommun &= matieres[s[0]].TD
        else:
            TDEnCommun &= matieres[seance[0]].TD

    if len(TDEnCommun) == 0:
        _logger.warning("Maquette : %s : %s   ==>   Aucun groupe en commun pour les matières de cette dépendance", matieres[dependanceMaquette[0]].nom, dependanceMaquette[1])

    erreurTD = ""
    for TD in TDEnCommun:
        cours = getCours(seances,TD,matieres,dependanceMaquette)

        if len(cours) == 1:

            if condition == "avant":
                for autreNature,autreCours in matieres[cours[0].codeMatiere].coursParTD.items():
                    #on compare uniquement à la première séance de chaque nature
                    if cours[0].nature != autreNature and len(autreCours[TD]) > 0:
                        if not cours[0] < autreCours[TD][0]:
                            erreurTD += TD
                            break

            elif condition == "apres":
                for autreNature,autreCours in matieres[cours[0].codeMatiere].coursParTD.items():
                    #on compare uniquement à la dernière séance de chaque nature
                    if cours[0].nature != autreNature and len(autreCours[TD]) > 0:
                        if not autreCours[TD][-1] < cours[0]:
                            erreurTD += TD
                            break

        else:
            if not all(avant(cours[i],cours[i + 1]) for i in range(len(cours) - 1)):
                erreurTD += TD

    return erreurTD


def lecteurOdf_mcc_reader(nomFichierOdf,naturesPossibles,listeTD):
    """
    Lit l'odf grâce à mcc_reader pour créer le dictionnaire des matières 
    """
    odf = mcc_reader.read_mcc_file(nomFichierOdf)

    matieres = {}
    for numeroSemestre,semestreOdf in enumerate(odf.semestres):
        for ue in semestreOdf.ues:
            for matiere in ue.matieres:
                mat = Matiere(code = matiere.code, nom = matiere.nom, TD = set(listeTD), semestre = numeroSemestre + 1)
                mat.coursParTD = {nature : {TD : [] for TD in mat.TD} for nature in naturesPossibles}
                mat.coursParTD["c"]["ODF"] = matiere.nb_seances_cm
                mat.coursParTD["ctd"]["ODF"] = matiere.nb_seances_ctd
                mat.coursParTD["td"]["ODF"] = matiere.nb_seances_td
                mat.coursParTD["tp"]["ODF"] = matiere.nb_seances_tp + matiere.nb_seances_tp2groupes
                mat.coursParTD["pr"]["ODF"] = matiere.nb_seances_projets
                matieres[mat.code] = mat
    
    return matieres


def lecteurOdf_xlsx2odf(nomFichierOdf,naturesPossibles,listeTD):
    """
    Lit l'odf grâce à xlsx2odf pour créer le dictionnaire des matières 
    """
    odf = xlsx2odf.read_odf(nomFichierOdf)
    matieres = {}
    for departement in odf.values():
        for numeroSemestre,semestreOdf in enumerate(departement["content"].values()):
            for ue in semestreOdf["content"].values():
                
                if len(ue["parcours"]) == 0: 
                    parcours = set(listeTD)
                else:
                    parcours = ue["parcours"]
                
                if "C" in parcours:
                    parcours = listeTD

                elif "L" in parcours:
                    parcours.discard("L")
                    for TD in listeTD:
                        if "L" in TD:
                            parcours.add(TD)

                for codeMatiere, matiere in ue["content"].items():
                    if codeMatiere in matieres:
                        for TD in parcours - matieres[codeMatiere].TD:
                            for nature in naturesPossibles:
                                matieres[codeMatiere].coursParTD[nature][TD] = []
                        matieres[codeMatiere].TD |= parcours       

                    else:
                        mat = Matiere(code = codeMatiere, nom = matiere["name"], TD = set(parcours), semestre = numeroSemestre + 1)
                        mat.coursParTD = {nature : {TD : [] for TD in mat.TD} for nature in naturesPossibles}         
                        matieres[codeMatiere] = mat
    
    return matieres


def lecteurXlsx(nomFichierConfig, nomFichierEdt, pageDebutEdt, nomFichierMaquette, nomFichierInstructor, nomFichierAlias, nomFichierOdf, lecteurOdf, nomFichierCsv, NomFichierDependance, logLevel = logging.WARNING):
    """
    Indique les erreurs dans un emploi du temps, crée un fichier csv contenant tous les cours de l'edt 
    et crée un fichier txt contenant toutes les dépendances
    Inputs :
        - nomFichierConfig (str) : le nom du fichier de config (xlsx_reader_config.toml)
        - nomFichierEdt (str) : le nom du fichier de l'emploi du temps
        - pageDebutEdt (int) : le numéro de la page (la numérotation commence à 1) à laquelle débute l'edt (après les onglets maquette et alias)
        - nomFichierMaquette (str) : le nom de la maquette (cela peut être le même que celui de l'edt)
        - nomFichierInstructor (str) : le nom du fichier avec la liste des enseignants (instructor.csv)
        - nomFichierAlias (str) : le nom du fichier d'alias des profs (instructor-default-aliases.csv)
        - nomFichierOdf (str) : le nom du fichier de l'offre de formation
        - lecteurOdf (function) : le lecteur qui permet de lire l'odf (lecteurOdf_mcc_reader pour les 1SN ou  lecteurOdf_xlsx2odf pour les autres)
        - nomFichierCsv (str) : le nom souhaité du fichier csv contenant les cours
        - NomFichierDependance (str) : le nom souhaité du fichier de dépendances
        - logLevel (logging) : le niveau du logger souhaité (par défaut à WARNING)
    """


    logging.basicConfig(format='%(name)s:%(levelname)s: %(message)s', level=logLevel)

    with open(nomFichierConfig, "rb") as fichierconfig:
        config = pytomlpp.load(fichierconfig)
        naturesPossibles = config["naturesPossibles"]
        ignores = config["ignores"]["ignores"]
        TDPossible = config["TDPossible"]
        annee = config["infoEdt"]["annee"]
        semestre = config["infoEdt"]["semestre"]

    promotion = nomFichierEdt.split("/")[-1].split("edt")[-1][:3].upper()  #ex: 1SN
    listeTD = TDPossible[promotion]

    matieres = lecteurOdf(nomFichierOdf,naturesPossibles,listeTD)

    try:
        matieres, dependances = lecteurMaquette(nomFichierMaquette,naturesPossibles,listeTD,matieres)
    except MaquetteError:
        _logger.critical("impossible de lire la maquette : erreur sur le nom des colonnes")
        return None

    aliasProfs1 = edt_matieres.ade_enseignants(nomFichierInstructor)
    aliasProfs2 = lecteurAlias(nomFichierAlias)
    aliasProfs = aliasProfs1 | aliasProfs2

    for matiere in matieres.values():
        matiere.numeroModule = mini(matiere.acronymes).upper()
        for acronyme in matiere.acronymes:
            if acronyme in aliasProfs:
                aliasProfs.pop(acronyme)

    
    workbookEdt = openpyxl.load_workbook(nomFichierEdt)

    with open(nomFichierCsv, "w", newline='', encoding='utf8') as fichierCours:
        writer = csv.writer(fichierCours, delimiter = ",")
        writer.writerow(["promotion", "numeroModule", "nomLongModule", "apogee", "nature", "groupes", "jour", "semaine", "heure", "duree", "salles", "profs", "couleur", "note"])

        mots_incompris = set()
        dispositionJour = None

        for sheetname in workbookEdt.sheetnames[pageDebutEdt - 1 : ]:
            
            if "S2" in sheetname:
                semestre = 2

            if sheetname in ("1", "1_Vacances"):
                annee += 1

            if not "vacance" in sheetname.lower():
                page = workbookEdt[sheetname]
                mergedCells = page.merged_cells.ranges

                nouvelleDispositionJour = [cellule.value for cellule in page[6]]
                nouvelleDispositionJour = nouvelleDispositionJour[nouvelleDispositionJour.index("LUNDI") : nouvelleDispositionJour.index("VENDREDI") + 1]
                if nouvelleDispositionJour != dispositionJour:
                    dispositionJour = nouvelleDispositionJour
                    colonneTD, colonneJour = generateurColonne(dispositionJour, listeTD)

                    for cellule in page["A"]:
                        if cellule.value == "8h-9h":
                            ligneDebutEdt = cellule.row
                            break

                for mergedCell in mergedCells:
                    coordonneesCellule = str(mergedCell).split(":")[0]
                    if coordonneesCellule not in ("A1","E4","J4","K4"):

                        cours = Cours(str(mergedCell))
                        
                        texteCellule = page[coordonneesCellule].value
                        if not isinstance(texteCellule,str):
                            continue
                        texteCellule = texteCellule.replace(",","")

                        #récupérer ce qu'il y a entre crochets en tant que note
                        if "[" and "]" in texteCellule:
                            indice1 = texteCellule.index("[")
                            indice2 = texteCellule.index("]")
                            cours.note = texteCellule[indice1+1:indice2]
                            texteCellule = texteCellule.replace(texteCellule[indice1:indice2+1],"")


                        elementsCellule = unidecode(texteCellule.lower()).split()

                        try:
                            #test si les éléments dans la cellule sont connus
                            for taille in range(len(elementsCellule), 0, -1):
                                for indice in range(len(elementsCellule) - taille, -1, -1):
                                    if indice + taille <= len(elementsCellule):
                                        element = " ".join(mot for mot in (elementsCellule[indice : indice + taille]))

                                        reste = analyserElement(element, cours, ignores, naturesPossibles, matieres, aliasProfs, semestre, unidecode(page.title))
                                        if reste is None:
                                            for _ in range(taille):
                                                elementsCellule.pop(indice)

                            if len(elementsCellule) > 0:
                                _logger.error("Semaine : %s, cellule : %s     ==>    Elément incompris : %s", unidecode(page.title), mergedCell, elementsCellule)
                                mots_incompris |= set(elementsCellule)

                            else:
                                #finir la création du cours    
                                if cours.codeMatiere != "":
                                    
                                    #s'il n'y a pas de nature pour le cours car la nature est dans le nom de la matière (ex: bilan semestre 7)
                                    if cours.nature == "":
                                        for nature,acronymes in naturesPossibles.items():
                                            for acronyme in acronymes:
                                                if acronyme in matieres[cours.codeMatiere].nom.split(" "):
                                                    cours.nature = nature
                                    
                                    if cours.nature == "":
                                        _logger.error("Semaine : %s, cellule : %s     ==>    Il n'y a pas de nature pour le cours %s", unidecode(page.title), mergedCell, matieres[cours.codeMatiere].nom)

                                    else:
                                        #récupérer la semaine
                                        for chiffre in unidecode(page.title):
                                            if chiffre.isdigit():
                                                cours.semaine += chiffre
                                            else:
                                                break
                                        cours.semaine = int(cours.semaine)


                                        #récupérer la première colonne du cours
                                        colonneCoursDebut = "".join(lettre for lettre in str(mergedCell).split(":")[0] if not lettre.isdigit())
                                        colonneCoursDebutNumero = ord(colonneCoursDebut[-1]) - 64 + 26 * (len(colonneCoursDebut) - 1) #marche car on ne dépasse pas la colonne AZ
                                        
                                        #récupérer la dernière colonne du cours
                                        colonneCoursFin = "".join(lettre for lettre in str(mergedCell).split(":")[1] if not lettre.isdigit())
                                        colonneCoursFinNumero = ord(colonneCoursFin[-1]) - 64 + 26 * (len(colonneCoursFin) - 1) #marche car on ne dépasse pas la colonne AZ


                                        #récupérer le jour
                                        jour = [jour for jour,colonnes in colonneJour.items() if colonneCoursDebutNumero in colonnes][0]
                                        cours.jour = dateSemaine(annee, cours.semaine, jour)


                                        #récupérer l'heure
                                        ligneCoursDebut = int("".join(chiffre for chiffre in str(mergedCell).split(":")[0] if chiffre.isdigit()))
                                        ligneCoursFin = int("".join(chiffre for chiffre in str(mergedCell).split(":")[1] if chiffre.isdigit()))

                                        l = ligneDebutEdt
                                        dictLigneHeure = {(l,l) : ("8:00", 60), (l+1,l+1) : ("9:00", 60), (l+2,l+2) : ("10:15", 60),
                                                        (l+3,l+3) : ("11:00", 60), (l+5,l+5) : ("14:00", 60), (l+6,l+6) : ("15:00", 60),
                                                        (l+7,l+7) : ("16:15", 60), (l+8,l+8) : ("17:00", 60), (l+9,l+9) : ("18:00", 60),
                                                        (l,l+1) : ("8:00", 105), (l+2,l+3) : ("10:15", 105), (l+5,l+6) : ("14:00", 105),
                                                        (l+7,l+8) : ("16:15", 105), (l,l+3) : ("8:00", 240), (l+5,l+8) : ("14:00", 240)}
                                        cours.heure = dictLigneHeure[(ligneCoursDebut,ligneCoursFin)][0]
                                        cours.duree = dictLigneHeure[(ligneCoursDebut,ligneCoursFin)][1]


                                        #récupérer le groupe
                                        TDDebut = [TD for TD,colonnes in colonneTD.items() if colonneCoursDebutNumero in colonnes][0]
                                        TDFin = [TD for TD,colonnes in colonneTD.items() if colonneCoursFinNumero in colonnes][0]
                                        
                                        if cours.verificationGroupe != 0:
                                            if cours.verificationGroupe != listeTD.index(TDDebut) + 1:
                                                _logger.warning("Semaine : %s, cellule : %s     ==>    Le groupe de TD renseigné est incohérent avec la colonne", unidecode(page.title), mergedCell)
                                        
                                        
                                        for numeroTD in range(listeTD.index(TDDebut), listeTD.index(TDFin) + 1):
                                            TD = listeTD[numeroTD]
                                            cours.groupe += TD
                                            if TD in matieres[cours.codeMatiere].TD:
                                                matieres[cours.codeMatiere].coursParTD[cours.nature][TD].append(cours)
                                            else:
                                                _logger.error("Semaine : %s, cellule : %s     ==>    Le groupe %s n'est pas dans les parcours de la matière %s (parcours : %s)", unidecode(page.title), mergedCell, TD, matieres[cours.codeMatiere].nom, matieres[cours.codeMatiere].TD)   
                                                
                
                                        nomLongModule = matieres[cours.codeMatiere].nom
                                        numeroModule = matieres[cours.codeMatiere].numeroModule

                                        #écrit un cours dans le fichier csv
                                        writer.writerow([promotion, numeroModule, nomLongModule] + cours.csv())



                        except NatureError:
                            _logger.error("Semaine : %s, cellule : %s     ==>    Plusieurs natures pour le même cours", unidecode(page.title), mergedCell)
                            

    workbookEdt.close()

    _logger.warning("Mots imcopris : %s", mots_incompris)



    for matiere in matieres.values():
        for nature,cours in matiere.coursParTD.items():
            seancesParcourus = []
            for TD,seances in cours.items():
                if TD in matiere.TD:
                    seances.sort()

                    #vérification du numéro de la séance
                    for numero,seance in enumerate(seances):
                        seance.numero = numero + 1
                        if not seance.numeroEdt in (seance.numero, 0):  #0 si le numéro de la séance n'est pas préciser
                            if seance not in seancesParcourus:
                                seancesParcourus.append(seance)
                                _logger.warning("Semaine : %s, cellule : %s     ==>    Le numéro de %s %s groupes %s devrait être %s au lieu de %s", seance.semaine, seance.cellule, nature, matiere.nom, seance.groupe, seance.numero, seance.numeroEdt)


    for matiere in matieres.values():
        for nature,cours in matiere.coursParTD.items():
            nombreSeances = {}
            nombreOdf = None
            for TD,seances in cours.items():
                if TD in listeTD:
                    if TD in matiere.TD:
                        nombreSeances[TD] = len(seances)
                elif TD == "ODF":
                    nombreOdf = seances
                
            #Vérifier si tous les TDs ont le même nombre de séances
            nombreSeancesList = list(nombreSeances.values())
            if not all(nombreSeancesList[i] == nombreSeancesList[i+1] for i in range(len(nombreSeancesList) - 1)):
                _logger.error("%s %s   ==>   Nombre de séances par groupes : %s", nature, matiere.nom, nombreSeances)

            else: #Tous les TDs ont le même nombre de séances et on vérifie que c'est le même celui dans l'odf
                if nombreOdf != None and len(nombreSeancesList) > 0 and nombreSeancesList[0] != nombreOdf:
                    _logger.error("%s %s   ==>   %s séances dans l'edt au lieu de %s séances dans l'odf", nature, matiere.nom, nombreSeancesList[0], nombreOdf)



    for indice,dependance in enumerate(dependances):
        dependancesFormatee = lecteurDependance(dependance[0],dependance[1],matieres,naturesPossibles)
        try:
            if dependancesFormatee != None:
                if len(dependancesFormatee) == 1:
                    condition = dependancesFormatee[0].pop(3)
                    erreurTD = comparaison(dependancesFormatee,matieres,listeTD,dependances[indice],condition)
                    if len(erreurTD) != 0:
                        _logger.error("%s : %s   ==>   Cette dépendance n'est pas vérifiée pour les groupes %s", matieres[dependances[indice][0]].nom, dependances[indice][1], erreurTD)

                else:
                    erreurTD = comparaison(dependancesFormatee,matieres,listeTD,dependances[indice])
                    if len(erreurTD) != 0:
                        _logger.error("%s : %s   ==>   Cette dépendance n'est pas vérifiée pour les groupes %s", matieres[dependances[indice][0]].nom, dependances[indice][1], erreurTD)

        except IndexError:
            pass


    
    ecritureDependance(NomFichierDependance,matieres)




if __name__ == "__main__":
    nomFichierConfig = "xls2ade/src/xlsx_reader_config.toml"
    nomFichierEdt = "xls2ade/xls/src/edt1sn-2024.xlsx"
    pageDebut = 4
    nomFichierMaquette = "xls2ade/xls/src/edt1sn-2024-maquette.xlsx"    
    nomFichierInstructor = "xls2ade/src/instructor.csv"
    nomFichierAlias = "xls2ade/src/instructor-default-aliases.csv"
    nomFichierOdf = "xls2ade/xls/src/odf1sn-2024.xlsx"
    nomFichierCsv = "cours1sn.csv"
    lecteurOdf = lecteurOdf_mcc_reader
    nomFichierDependance = "dependances1sn.txt"
    lecteurXlsx(nomFichierConfig, nomFichierEdt, pageDebut, nomFichierMaquette, nomFichierInstructor, nomFichierAlias, nomFichierOdf, lecteurOdf, nomFichierCsv, nomFichierDependance, logLevel=logging.WARNING)


